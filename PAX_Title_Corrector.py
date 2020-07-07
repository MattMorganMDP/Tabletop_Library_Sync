import csv
import openpyxl
from difflib import get_close_matches
import sys 
from time import sleep 
from datetime import datetime
from os import path

#All code contained within main() function as to prevent automatic execution when this file is imported into main menu script

def main():
    
    ########################################################################
    # Function to check if a game name has multiple entries on BGG
    ########################################################################

    def DupesCheck(name):
        
        #See if the dictionary value for game name is a nested dictionary with more than one entry. If so, print them to menu and prompt user to choose one.
        if len(BGGnames[name]) > 1:
            #for games in that dict, enumerate their values and print to screen
            for index, game in enumerate(BGGnames[name].keys()):
                print(str(index + 1) + ' - BGG ID#: ' + str(game) + ' - Year Published: ' + str(BGGnames[name][game]))    
            
            # Take user input to select among duplicates, while taking measures to validate input            
            selected_ID = 0 
            while True:
                #Ensure an integer is input
                try: 
                    selected_ID = int(input('Please enter the index # of the selected match: '))  #could add a reference dict in the enumerate line above, with index as the key, BGG ID as the value
                except ValueError:
                    print('Input must be numeric')
                    continue

                #Then ensure that integer is within range (length of the nested dictionary)
                if (int(selected_ID) > len(BGGnames[name].keys()) or int(selected_ID <= 0)):  
                    print ('Sorry, input is out of range')
                    continue
                else:
                    break

            lock = True  #Set flag locking row from future Corrector runs, so user does not have to re-resolve dupes
           
        else:
            selected_ID = list(BGGnames[name].keys())[0]
            lock = False  #Do not set flag locking row. There was only 1 BGG match, so there is no burden to user in including this entry in future re-runs of script.
            #There is a higher likelihood, albeit small one, that the game is not yet listed on BGG, but is being matched to a classic title of same name

        return selected_ID, lock
    
    ########################################################################
    # Function to write a row to the PAX Corrections csv
    ########################################################################
   
    def TitleWriting(game, newname = '.', status = 'pass'):
        
        # Test for correction conditions . 
        # 1) Newname == . means no newname value was passed (a direct match with BGG was present), so write row as-is and add the BGG ID#
        # 2) A blank newname means correction was attempted but failed, so write row with a 0 in BGG ID# field. Log error.
        # 3) Status flag set to fail means game was manual renamed. Need to write newname, but not attempt to find BGG ID#. Log error. 
        # 4) If not blank or ., and status flag not changed, write the newname variable as game title and look up its BGG ID#
        if newname == '.':
            ID, lock = DupesCheck(game)
            TitleWriter.writerow([game, int(PAXids[PAXnames.index(game)]), ID, lock])
        elif newname == '':
            TitleWriter.writerow([game, int(PAXids[PAXnames.index(game)]), 0, True])
            ErrorWriter.write(game + '\n')
        elif status == 'fail':
            TitleWriter.writerow([newname, int(PAXids[PAXnames.index(game)]), 0, True])
            ErrorWriter.write(newname + '\n')
        else:
            ID, lock = DupesCheck(newname)
            TitleWriter.writerow([newname, int(PAXids[PAXnames.index(game)]), ID, lock])
        return()

    ########################################################################
    # Function to display list of matches and take user input
    ########################################################################

    def MatchSelector(match, title):
        #Function takes in the PAX title and its corresponding get_close_matches object

        print(title + ' - Has potential correction from BGG search:')
        for index, name in enumerate(match):
            print(index + 1, name) #index is incremented for user-friendliness, so list does not start at 0
        print(len(match) + 1, 'No match') #Add an extra entry, one above match list's range, for "N/A" option
        #TO DO - Work around instances where there are multiple identical options. Pull in BGG metadata?

        ### Take user input to select match            
        selected = 0 
        while True:
            #Ensure an integer is input
            try: 
                selected = int(input('Please select a match: '))
            except ValueError:
                print('Input must be numeric')
                continue

            #Then ensure that integer is within range (match index plus 2: 1 so that it doesn't start at 0, plus another 1 to account for "N/A" option)
            if (int(selected) not in range (1,len(match) + 2)):  
                print ('Sorry, input is out of range')
                continue
            else:
                break
        
        return selected

    ###################################################################################################################################
    # Function to drive attempting of a match. Calls MatchSelector and TitleWriting functions. Can sense if first or subsequent attempt.
    ###################################################################################################################################

    def AttemptMatch(game, BGGnames, manual_name = ''):
        if manual_name == '':
            match = get_close_matches(game, BGGnames.keys(), n=3, cutoff = 0.7)
        else:
            match = get_close_matches(manual_name, BGGnames.keys(), n=3, cutoff = 0.7)

        if len(match) != 0:
            selected = MatchSelector(match,game)

            ### Check if a match was selected, or if user chose N/A "No match" option, which will not be a valid index of match list    
            if int(selected) <= int(len(match)):
                print('Thank you for selecting input #' + str(selected) +": " + match[int(selected) - 1])
                TitleWriting(game, match[int(selected) - 1])
                return 'success'
            else:
                return 'fail'  #no match selected
        else:
            return 'fail' #no matches found

    ########################################################################
    # Main body code
    ########################################################################

    ###### Prepare file I/O: Load Tabletop Library titles report csv ######

    # Initialize lists
    PAXnames = []
    PAXids = []
    PAXstatus = []

    # Present user with menu selection of frequently-used .csv files, or allow to input own filename
    print('\n') 
    print (30 * '-')
    print ("   D A T A - I N P U T")
    print (30 * '-')
    print ("1. Tabletop Library's Titles Report (TTLibrary_Titles.csv)")
    print ("2. Prior Corrections Script Output (PAXcorrections.csv)")
    print ("3. Prior Export (Manual Filename Entry)")
    print ("4. Return to Main Menu")
    print (30 * '-')
    
    # Prompt user for input and validate input
    choice = 0 
    while True:
        #Ensure an integer is input
        try: 
            choice = int(input('Please enter your choice [1-4]: '))
        except ValueError:
            print('Input must be numeric')
            continue

        #Then ensure that integer is within range (1-4)
        if (int(choice) > 4 or int(choice <= 0)):  
            print ('Sorry, input is out of range')
            continue
        else:
            break

    # Branching logic for menu choices
    # TO-DO: Once Tabletop Library is updated to include a corrections report (which will match output format of this script), tree can be simplified
    if choice == 1:
        try:
            PAXgames = open('TTLibrary_Titles-testshort.csv', mode='r', newline='')
            reader = csv.reader(PAXgames) # Read .csv into a variable
            next(reader)
            for rows in reader:  # Iterate through input csv and append different elements of each row to the appropriate list
                PAXnames.append(rows[0])
                PAXids.append(rows[4])
        except:
            print('Error: TTLibrary_Titles.csv not found in current working directory. Returning to main menu...')
            sleep(2)
            return
    elif choice == 2:
        try:
            PAXgames = open('PAXcorrections.csv', 'r', newline='', encoding='utf-16')
            reader = csv.reader(PAXgames) # Read .csv into a variable
            header = next(reader)
            for rows in reader:  # Iterate through input csv and append different elements of each row to the appropriate list
                PAXnames.append(rows[0])
                PAXids.append(rows[1])
                PAXstatus.append(rows[2])
        except:
            print('Error: PAXcorrections.csv not found in current working directory. Returning to main menu...')
            sleep(2)
            return
    elif choice == 3:
        PAX_Titles_path = input("Enter the name of past script output .csv file: ")
        try:
            PAXgames = open(PAX_Titles_path, 'r', newline='')
            reader = csv.reader(PAXgames) # Read .csv into a variable
            header = next(reader)
            for rows in reader:  # Iterate through input csv and append different elements of each row to the appropriate list
                PAXnames.append(rows[0])
                PAXids.append(rows[1])
                PAXstatus.append(rows[2])
        except:
            print ('Could not load file successfully. Returning to main menu...')
            sleep(2)
            return

    elif choice == 4:
        print('Returning to main menu...')
        sleep(2)
        return
 
    # TO-DO: Once PAX Correction report is generated by Tabletop Library, will no longer need to overwrite header here
    header = ['Title', 'PAX ID', 'BGG ID', 'Status']

    
    ###### Prepare file I/O: Load BGG names: Open BGG workbook, set active sheet ######
    print('Loading BGG ID index table. This will take some time...')
    try:
        wb = openpyxl.load_workbook('BGG_IDs.xlsx')
    except: 
        print('Error: Could not locate index of BGG names/IDs. \n Please load BGG_ID_spreadsheet_complete.xlsx into current working directory and restart script')
        return
    sheet = wb.active

    # Initialize a dictionary of BGG names. Each unique name is a key. Value is a dictionary of ID#/year published pairings
    # Value dictionary varies in length depending on how many duplicate entries there are on BGG for that game title
    BGGnames = {}
    for x in range(1,sheet.max_row + 1):
        if str(sheet.cell(x,2).value) in BGGnames.keys():
            BGGnames[str(sheet.cell(x,2).value)].update({sheet.cell(x,1).value : sheet.cell(x,3).value})
        else:
            BGGnames[str(sheet.cell(x,2).value)] = {sheet.cell(x,1).value : sheet.cell(x,3).value}

    ###### Prepare file I/O: Open files to be written ######

    #Open output csv for writing corrected titles, set the writer object, and write the header
    PAXcorrections = open('PAXcorrections.csv', 'w', newline='', encoding='utf-16')
    TitleWriter = csv.writer(PAXcorrections, delimiter=',', escapechar='\\', quotechar='"', quoting=csv.QUOTE_NONNUMERIC)
    TitleWriter.writerow(header)

    #Open a text file for mismatch error logging. Write header message with timestamp.
    ErrorWriter = open('Title_Mismatch_Log.txt', 'a+') #append mode used to continue growing log
    ErrorWriter.write('\n \nTitle Corrector was run on ' + str(datetime.now()) + '\nThe following games could not be linked to BGG ID#s: \n')

    ###### Compare game names and write rows to excel sheet ######
    for game in PAXnames:
        if game not in BGGnames.keys():
            print('Attempting match of ' + game)
            status = AttemptMatch(game, BGGnames)
            
            # If match was not found, try again with manually-input revision to title.
            if status == 'fail':
                print('No match was found on first attempt')
                manual_title = input('Please manually enter a corrected title. If unsure, leave blank to skip: ')
                if manual_title == '':
                    TitleWriting(game, manual_title)
                    status = 'done'  #revert the status code returned by AttemptMatch() so user is not prompted for manual input again in next code block (if status == 'fail')
                else:
                    status = AttemptMatch(game, BGGnames, manual_title)
                    
            if status == 'fail':
            # If match was still not found on second attempt, allow for full manual correction without BGG link
                print(game + ' - Has no BGG match found, and was unable to be corrected through manual adjustment')
                manual_title = input('Please manually enter a corrected title. This will be written to the PAX dB without attempt to find a BGG match. If unsure, leave blank to skip: ')
                TitleWriting(game, manual_title, 'fail')
        
        # Clear match was found between PAX & BGG:
        else:
            print(game +  ' - No correction needed. BGG ID# is ' + str(list(BGGnames[game].keys())[0]) + ' PAX ID# is ' + str(PAXids[PAXnames.index(game)]))
            TitleWriting(game)

        print('\n') 

    PAXcorrections.close()
    ErrorWriter.close()

if __name__ == "__main__":
    main()